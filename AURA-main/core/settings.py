import os
import json
import logging
from logging.handlers import RotatingFileHandler
from openpyxl.styles import PatternFill

class Constants:
    ROOT_DIR               = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    LOG_DIR                = os.path.join(ROOT_DIR, "logs")
    TEST_FLOWS_PATH        = os.path.join(ROOT_DIR, "test_flows")
    VALIDATE_OUTPUT_DIR    = os.path.join(ROOT_DIR, "test_flows_exec_results")
    SCRAPE_OUTPUT_DIR      = os.path.join(ROOT_DIR, "z_scraped_data")
    LOCALE_MAP_PATH        = os.path.join(ROOT_DIR, "test_flows_common", "locale_map.json")
    ADDRESS_DATA_PATH      = os.path.join(ROOT_DIR, "test_flows_common", "address_data.json")
    PRINTER_PROFILES_PATH  = os.path.join(ROOT_DIR, "test_flows_common", "printer_profiles.json")
    SCRAPE_HEADERS = [
        "Context", "SubContext", "PageURL", "Timestamp", "Type", "Tag", "Text", "Href", "Screenshot", "ExcludeRegions"
    ]
    VALIDATION_HEADERS = [
        "Context", "SubContext", "PageURL", "Timestamp", "Type", "Tag",
        "Text_Expected", "Text_Actual",
        "Href_Expected", "Href_Actual",
        "Text_Result", "Screenshot_Baseline", "Screenshot_Validation",
        "Screenshot_Result", "Diff_Image", "ExcludeRegions", "SuggestedExcludeRegions", "Comments"
    ]
    COLOR_HEADER                    = "A67657"
    COLOR_FAIL                      = "FF9999"
    COLOR_PASS                      = "CCFFCC"
    COLOR_PASS_WITH_EXCLUDES        = "FF7518"
    COLOR_PASS_TEXT_ONLY            = "CCCCFF"
    CELL_FILL_FAIL                  = PatternFill(start_color=COLOR_FAIL,   end_color=COLOR_FAIL,   fill_type="solid")
    CELL_FILL_PASS                  = PatternFill(start_color=COLOR_PASS,   end_color=COLOR_PASS,   fill_type="solid")
    CELL_FILL_PASS_WITH_EXCLUDES    = PatternFill(start_color=COLOR_PASS_WITH_EXCLUDES, end_color=COLOR_PASS_WITH_EXCLUDES, fill_type="solid")
    CELL_FILL_PASS_TEXT_ONLY        = PatternFill(start_color=COLOR_PASS_TEXT_ONLY, end_color=COLOR_PASS_TEXT_ONLY, fill_type="solid")
    IGNORE_TAGS                  = {
        "script", "style", "noscript", "template", "meta", "link", "head", "title",
        "svg", "defs", "symbol", "path", "clippath"
    }
    DEFAULT_TIMEOUT        = 30 * 1000
    MAX_RETRIES            = 2

# Logging
os.makedirs(Constants.LOG_DIR, exist_ok=True)
_LOG_LEVEL = logging.INFO  # Default; overridden in aura.py if --debug

def set_framework_log_level(level):
    global _LOG_LEVEL
    _LOG_LEVEL = level
    framework_logger.setLevel(level)
    for handler in framework_logger.handlers:
        handler.setLevel(level)

def _setup_framework_logger(
    name: str = "AURA",
    log_file: str = "framework.log",
    max_bytes: int = 5 * 1024 * 1024,
    backup_count: int = 5,
) -> logging.Logger:
    logger = logging.getLogger(name)
    logger.setLevel(_LOG_LEVEL)
    logger.propagate = False

    if not logger.handlers:
        formatter = logging.Formatter('[%(asctime)s] [%(levelname)s] %(message)s')
        file_path = os.path.join(Constants.LOG_DIR, log_file)
        file_handler = RotatingFileHandler(
            filename=file_path,
            maxBytes=max_bytes,
            backupCount=backup_count,
            encoding='utf-8'
        )
        file_handler.setLevel(_LOG_LEVEL)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

        console_handler = logging.StreamHandler()
        console_handler.setLevel(_LOG_LEVEL)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
    return logger

framework_logger = _setup_framework_logger()

class GlobalState:
    stack                   = None
    locale                  = None
    language                = None
    printer_profile         = None
    easy_enroll             = None
    biz_model               = None
    target                  = None
    region                  = None
    country                 = None
    country_code            = None
    language_code           = None
    address_payload         = None
    selected_printer_sku    = None
    tenant_email            = None
    flow_name               = None
    timestamp               = None
    mode                    = None
    run_dir                 = None
    headless                = False
    payment_method          = None
    simulator_platform      = None
    collect_logs            = False
    requirements            = {}
    
    @classmethod
    def reset(cls):
        cls.stack                   = None
        cls.locale                  = None
        cls.language                = None
        cls.printer_profile         = None
        cls.easy_enroll             = None
        cls.biz_model               = None
        cls.region                  = None
        cls.country                 = None
        cls.country_code            = None
        cls.language_code           = None
        cls.address_payload         = None
        cls.selected_printer_sku    = None
        cls.tenant_email            = None
        cls.flow_name               = None
        cls.timestamp               = None
        cls.mode                    = None
        cls.run_dir                 = None
        cls.headless                = False
        cls.target                  = None
        cls.payment_method          = None
        cls.simulator_platform      = None
        cls.collect_logs            = False
        cls.requirements            = {}

headless_on = lambda: GlobalState.headless

class Config:
    @classmethod
    def init(cls, **kwargs):
        GlobalState.reset()
        GlobalState.stack           = kwargs.get("stack")
        GlobalState.locale          = kwargs.get("locale")
        GlobalState.language        = kwargs.get("language")
        GlobalState.printer_profile = kwargs.get("profile")
        GlobalState.easy_enroll     = kwargs.get("easy_enroll")
        GlobalState.biz_model       = kwargs.get("biz_model", "Flex")
        GlobalState.headless        = kwargs.get("headless", False)
        GlobalState.mode            = kwargs.get("mode")
        GlobalState.flow_name       = kwargs.get("flow_name")
        GlobalState.timestamp       = kwargs.get("timestamp")
        GlobalState.run_dir         = kwargs.get("run_dir")
        GlobalState.target          = kwargs.get("target")
        GlobalState.payment_method  = kwargs.get("payment_method", "credit_card_master")
        GlobalState.simulator_platform = kwargs.get("simulator_platform")
        GlobalState.requirements    = kwargs.get("requirements", {})
        cls._load_locale_info()
        cls._load_address_data()
        cls._resolve_printer_sku()
        for k, v in kwargs.items():
            setattr(GlobalState, k, v)

    @staticmethod
    def _load_locale_info():
        with open(Constants.LOCALE_MAP_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if GlobalState.locale not in data:
            raise ValueError(f"Locale '{GlobalState.locale}' not found in locale_map.json")
        locale_data = data[GlobalState.locale]
        GlobalState.country       = locale_data["country"]
        GlobalState.country_code  = locale_data["code"]
        GlobalState.region        = locale_data["region"]
        languages                = locale_data["languages"]
        GlobalState.language      = GlobalState.language or languages[0]
        GlobalState.language_code = f"{GlobalState.language}-{GlobalState.country_code}"

    @staticmethod
    def _load_address_data():
        with open(Constants.ADDRESS_DATA_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if GlobalState.country not in data:
            raise ValueError(f"No address found for country: {GlobalState.country}")
        GlobalState.address_payload = data[GlobalState.country]

    @staticmethod
    def _resolve_printer_sku():
        if not GlobalState.printer_profile:
            return
        with open(Constants.PRINTER_PROFILES_PATH, "r", encoding="utf-8") as f:
            printers = json.load(f)
        profile = GlobalState.printer_profile
        if profile not in printers:
            raise ValueError(f"Printer profile '{profile}' not found")
        skus = printers[profile]["skus"].get(GlobalState.region, [])
        if not skus:
            raise ValueError(f"No SKUs for region '{GlobalState.region}' in profile '{profile}'")
        if GlobalState.easy_enroll is not None:
            desired = str(GlobalState.easy_enroll).lower() == "yes"
            skus = [s for s in skus if s.get("easy_enroll") == desired]
            if not skus:
                raise ValueError(f"No SKUs with easy_enroll={desired} in profile '{profile}'")
        GlobalState.selected_printer_sku = skus[0]["sku"]
