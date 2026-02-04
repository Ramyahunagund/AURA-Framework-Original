# core/utils.py

import os
import sys
import re
import json
import string
import random
import hashlib
import importlib
import time
import io
import tempfile
from datetime import datetime
import traceback
from PIL import Image, ImageChops
from typing import Callable, Optional, Dict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, ElementHandle
from core.settings import Constants, Config, framework_logger
import zipfile
from pymongo import MongoClient
import boto3
from botocore.exceptions import ClientError

# =========================== Basic Utilities ===========================

def get_current_timestamp() -> str:
    return datetime.now().strftime("%Y_%m_%d_%H_%M_%S")

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name)

def generate_random_string(length: int = 6) -> str:
    pool = string.ascii_lowercase + string.digits
    return ''.join(random.choices(pool, k=length))

def hash_text(text):
    if isinstance(text, bytes):
        to_hash = text
    else:
        to_hash = (text or "").encode("utf-8")
    return hashlib.sha256(to_hash).hexdigest()[:12]


def ensure_dirs(*dirs: str) -> None:
    for d in dirs:
        os.makedirs(d, exist_ok=True)

class ElementsDict(dict):
    def __getattr__(self, name):
        try:
            return self[name]
        except KeyError:
            raise AttributeError(f"'ElementsDict' object has no attribute '{name}'")

    def __setattr__(self, name, value):
        self[name] = value

# ===================== Excel Loader and Sheet Utilities =====================

def get_or_create_sheet(workbook, sheet_name: str, headers: list = None):
    """
    Return the sheet by name, or create it (with headers) if missing.
    """
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    sheet = workbook.create_sheet(title=sheet_name)
    hdrs = headers if headers is not None else Constants.SCRAPE_HEADERS
    for idx, header in enumerate(hdrs, start=1):
        cell = sheet.cell(row=1, column=idx, value=header)
        cell.alignment = Alignment(horizontal="center")
    return sheet

def find_row(sheet, selector: str) -> bool:
    """
    Return True if any value in the first column matches 'selector'.
    """
    for (cell_value,) in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if cell_value == selector:
            return True
    return False

def append_scraped_row(sheet, row_data: list):
    """
    Append row_data to the sheet if not already present (by selector).
    """
    selector = row_data[0]
    if not selector or find_row(sheet, selector):
        framework_logger.debug(f"Skipping duplicate or empty selector row: {selector}")
        return
    sheet.append(row_data)

def auto_adjust_column_widths(sheet):
    """
    Resize columns based on maximum content length for readability.
    """
    for column_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        sheet.column_dimensions[col_letter].width = max_len + 2

def get_sheet_by_name(workbook, sheet_name: str):
    """
    Return sheet by name or None (with a warning).
    """
    if workbook and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    framework_logger.warning(f"Sheet '{sheet_name}' not found in workbook.")
    return None

def load_or_create_workbook(path: str, headers: list):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        ws = None
    return wb, ws

def save_workbook(wb, path: str):
    dirpath = os.path.dirname(path)
    if dirpath:
        os.makedirs(dirpath, exist_ok=True)
    wb.save(path)

def get_excel_contexts(excel_path: str) -> list:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    contexts = set()
    for flow, context, *_ in ws.iter_rows(min_row=2, values_only=True):
        if flow and context:
            contexts.add(context)
    return sorted(contexts)

# =========================== Playwright Utilities ===========================

def is_element_hidden_or_disabled(elem: ElementHandle) -> bool:
    return elem.evaluate("el => el.offsetParent === null || el.disabled === true")

def get_element_position(elem: ElementHandle) -> dict:
    box = elem.bounding_box() or {}
    return {
        "x": int(box.get("x", 0)),
        "y": int(box.get("y", 0)),
        "width": int(box.get("width", 0)),
        "height": int(box.get("height", 0))
    }

def extract_styles(elem: ElementHandle) -> dict:
    return elem.evaluate('''
        el => {
            const s = window.getComputedStyle(el);
            return {
                font_family: s.fontFamily || '',
                font_size: s.fontSize || '',
                font_weight: s.fontWeight || '',
                color: s.color || '',
                background: s.backgroundColor || '',
                text_align: s.textAlign || ''
            };
        }
    ''')

def extract_text(elem: ElementHandle) -> str:
    return elem.inner_text().strip()

def freeze_scroll(page: Page):
    page.add_style_tag(content="html, body { overflow: hidden !important; }")

def unfreeze_scroll(page: Page):
    page.add_style_tag(content="html, body { overflow: auto !important; }")

def take_element_screenshot(page: Page, selector: str, output_path: str) -> bool:
    try:
        element = page.locator(selector)
        if element.count() == 0 or not element.is_visible():
            return False
        dirpath = os.path.dirname(output_path)
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)
        element.screenshot(path=output_path)
        return True
    except Exception as e:
        framework_logger.warning(f"Could not screenshot '{selector}': {e}")
        return False

def take_element_handle_screenshot(handle: ElementHandle, output_path: str) -> bool:
    try:
        if not handle.is_visible():
            return False
        dirpath = os.path.dirname(output_path)
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)
        handle.screenshot(path=output_path)
        return True
    except Exception as e:
        framework_logger.warning(f"Could not screenshot element: {e}")
        return False

def save_full_page_screenshot(page: Page, flow_name: str, context_name: str, output_dir: str = None) -> str:
    directory = output_dir or Constants.VALIDATE_OUTPUT_DIR
    os.makedirs(directory, exist_ok=True)
    filename = sanitize_filename(f"{flow_name}_{context_name}_{get_current_timestamp()}.png")
    page.screenshot(path=os.path.join(directory, filename), full_page=True)
    return filename

# ======================= Excel/Validation Reporting =======================

def highlight_validation(ws, row_cells, passed: bool = None):
    header_row = [cell.value for cell in ws[1]]
    col_indices = {col: idx for idx, col in enumerate(header_row)}

    text_result_idx = col_indices.get("Text_Result")
    screenshot_result_idx = col_indices.get("Screenshot_Result")

    def norm(val):
        if val is None:
            return ""
        v = str(val).strip().upper()
        return v if v else ""

    text_val = norm(row_cells[text_result_idx].value) if text_result_idx is not None else ""
    screenshot_val = norm(row_cells[screenshot_result_idx].value) if screenshot_result_idx is not None else ""

    if text_val in ("", "N/A") and screenshot_val in ("", "N/A"):
        return
    if text_val == "PASS" and screenshot_val == "PASS":
        fill = Constants.CELL_FILL_PASS
    elif text_val in ("", "N/A") and screenshot_val == "PASS":
        fill = Constants.CELL_FILL_PASS
    elif text_val == "PASS" and screenshot_val in ("", "N/A"):
        fill = Constants.CELL_FILL_PASS_TEXT_ONLY
    elif text_val == "PASS" and screenshot_val == "PASS_WITH_EXCLUDES":
        fill = Constants.CELL_FILL_PASS_WITH_EXCLUDES
    elif text_val in ("", "N/A") and screenshot_val == "PASS_WITH_EXCLUDES":
        fill = Constants.CELL_FILL_PASS_WITH_EXCLUDES
    elif text_val == "FAIL" and screenshot_val == "FAIL":
        fill = Constants.CELL_FILL_FAIL
    elif text_val == "FAIL" and screenshot_val == "PASS":
        fill = Constants.CELL_FILL_FAIL
    elif text_val == "PASS" and screenshot_val == "FAIL":
        fill = Constants.CELL_FILL_FAIL
    else:
        fill = Constants.CELL_FILL_FAIL

    for cell in row_cells:
        cell.fill = fill

def generate_report(data: list, output_dir: str, filename: str = "validation_results"):
    os.makedirs(output_dir, exist_ok=True)
    json_path = os.path.join(output_dir, f"{filename}.json")
    html_path = os.path.join(output_dir, f"{filename}.html")
    with open(json_path, "w") as jf:
        json.dump(data, jf, indent=2)
    with open(html_path, "w") as hf:
        hf.write("<html><head><title>Validation Report</title></head><body>")
        hf.write("<h1>Validation Summary</h1><table border='1' cellpadding='5'>")
        if data:
            headers = list(data[0].keys())
            hf.write("<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>")
            for row in data:
                hf.write("<tr>" + "".join(f"<td>{row.get(h,'')}</td>" for h in headers) + "</tr>")
        hf.write("</table></body></html>")

def write_excel_context(rows, wb, sheet_name, path, header, highlight_func=None):
    from openpyxl.styles import PatternFill
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    if ws.max_row == 0:
        ws.append(header)
        header_fill = PatternFill(start_color=Constants.COLOR_HEADER, end_color=Constants.COLOR_HEADER, fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
    else:
        first_row = [cell.value for cell in ws[1]]
        if first_row != header:
            ws.delete_rows(1)
            ws.append(header)
            header_fill = PatternFill(start_color=Constants.COLOR_HEADER, end_color=Constants.COLOR_HEADER, fill_type="solid")
            for cell in ws[1]:
                cell.fill = header_fill
    rows_to_delete = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        values = [cell.value for cell in row]
        if all((v is None or str(v).strip() == "") for v in values):
            rows_to_delete.append(idx)
        elif values == header:
            rows_to_delete.append(idx)
    for idx in reversed(rows_to_delete):
        ws.delete_rows(idx)
    for row in rows:
        if not row or all((v is None or str(v).strip() == "") for v in (row.values() if isinstance(row, dict) else row)):
            continue
        if isinstance(row, dict):
            ws.append([row.get(col, "") for col in header])
        else:
            ws.append(row)
        if highlight_func is not None:
            highlight_func(ws, ws[ws.max_row], None)
    save_workbook(wb, path)

def write_report_excel_context(results, wb, sheet_name, path):
    write_excel_context(results, wb, sheet_name, path, Constants.VALIDATION_HEADERS, highlight_func=highlight_validation)

# ======================= Flow Runner Utilities =======================

def camel_to_snake(name: str) -> str:
    """
    Convert CamelCase flow names to snake_case module/function names.
    e.g. CreateV3Account -> create_v3_account
    """
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1_\2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1_\2', s1).lower()

def get_lang_code(flow_args: dict) -> str:
    return flow_args.get("tenant_country_language_code")

def make_target_directory(target="", simulator_platform=""):
    """
    Get the target directory name based on target and simulator_platform parameters from flow args.
    """
    target_folder_mapping = {
        "web_chrome": "chrome",
        "web_firefox": "firefox", 
        "web_safari": "safari",
        "GothamDesktop": "wingotham",
        "GothamMac": "macgotham",
        "Android": "android",
        "IOS": "ios",
        "HpxWindows": "hpxwindows",
        "HpxMac": "hpxmac",
        "HpxAndroid": "hpxandroid",
        "HpxIos": "hpxios",
        "hpSmartWinOnboardingAgent": "wingotham-od",
        "hpSmartMacOsOnboardingAgent": "macgotham-od",
        "hpSmartAndroidOnboardingAgent": "android-od",
        "hpSmartIosOnboardingAgent": "ios-od",
        "hpxwindows": "hpxwindows-od",
        "hpxmac": "hpxmac-od",
        "hpxandroid": "hpxandroid-od",
        "hpxios": "hpxios-od",
    }
    
    if target.startswith("web_"):
        return target_folder_mapping.get(target, target)
    else:
        return target_folder_mapping.get(simulator_platform, target)

def run_flow(
    mode: str,
    flow_name: str,
    flow_args: Optional[Dict] = None,
    context_override: Optional[str] = None
) -> None:
    if not flow_name:
        framework_logger.error("Flow name is required.")
        return

    flow_args = flow_args or {}
    module_basename = camel_to_snake(flow_name)
    module_path = f"test_flows.{flow_name}.{module_basename}"
    try:
        flow_module = importlib.import_module(module_path)
    except ModuleNotFoundError as e:
        framework_logger.error(f"Flow module not found: {module_path}")
        framework_logger.error(f"Error: {e}\n{traceback.format_exc()}")
        return
    if not hasattr(flow_module, module_basename):
        framework_logger.error(f"Expected function '{module_basename}' in '{module_path}'")
        return
    flow_func: Callable = getattr(flow_module, module_basename)
    run_id = get_current_timestamp()
    lang_code = get_lang_code(flow_args)
    excel_filename = f"{module_basename}_testdata.xlsx"
    screenshots_subdir = "screenshots"
    target = flow_args.get("target", "")
    simulator_platform = flow_args.get("simulator_platform", "")
    target_dir = make_target_directory(target, simulator_platform)
    collect_logs = flow_args.get("collect_logs")
    logs_enabled = bool(collect_logs)
    
    if mode == "scrape":
        output_dir = os.path.join(Constants.SCRAPE_OUTPUT_DIR, flow_name, run_id, lang_code, target_dir)
        os.makedirs(os.path.join(output_dir, screenshots_subdir), exist_ok=True)
        excel_path = os.path.join(output_dir, "scrape_data.xlsx")

        def stage_callback(context: str, page, screenshot_only=False, sub_context=None, animations=False):
            from core.scraper import scrape_context
            try:
                scrape_context(
                    flow_name,
                    context,
                    page,
                    excel_path,
                    output_dir,
                    screenshot_only=screenshot_only,
                    sub_context=sub_context,
                    lang_code=lang_code,
                    animations=animations,
                    target_dir=target_dir
                )
            except Exception as e:
                framework_logger.error(f"[SCRAPE] Error in context '{context}': {e}", exc_info=True)
 
    elif mode == "validate":
        output_dir = os.path.join(Constants.VALIDATE_OUTPUT_DIR, f"{flow_name}_{lang_code}_{target_dir}_{run_id}")
        screenshots_dir = os.path.join(output_dir, "validation_screenshots")
        failure_screens_dir = os.path.join(output_dir, "validation_failure_screenshots")
        os.makedirs(screenshots_dir, exist_ok=True)
        os.makedirs(failure_screens_dir, exist_ok=True)
        excel_path = os.path.join(Constants.TEST_FLOWS_PATH, flow_name, "test_data", lang_code, target_dir, excel_filename)
        def stage_callback(context: str, page, screenshot_only=False, sub_context=None, animations=False):
            from core.validator import validate_context
            try:
                validate_context(
                    flow_name,
                    context,
                    page,
                    excel_path,
                    output_dir,
                    screenshot_only=screenshot_only,
                    sub_context=sub_context,
                    lang_code=lang_code,
                    animations=animations,
                    target_dir=target_dir
                )
            except Exception as e:
                framework_logger.error(f"[VALIDATE] Error in context '{context}': {e}", exc_info=True)
 
    elif mode == "noop":
        output_dir = os.path.join(Constants.VALIDATE_OUTPUT_DIR, f"{flow_name}_{lang_code}_{target_dir}_{run_id}_noop")
        os.makedirs(output_dir, exist_ok=True)
        excel_path = None
        def stage_callback(context: str, page, screenshot_only=False, sub_context=None, animations=False):
            return
 
    else:
        framework_logger.error(f"Invalid mode: {mode!r}. Use 'scrape', 'validate', or 'noop'.")
        return
 
    config_kwargs = {
        **flow_args,
        "mode": mode,
        "flow_name": flow_name,
        "timestamp": run_id,
        "run_dir": output_dir,
        "headless": flow_args.get("headless", False),
        "collect_logs": collect_logs,
    }
    Config.init(**config_kwargs)
    framework_logger.info(f"Starting {mode} run for flow '{flow_name}' for {lang_code} (run_id={run_id})")
    log_type_map = {"har": "HAR only", "video": "Video only", "both": "HAR and Video"}
    if logs_enabled:
        framework_logger.info(f"{log_type_map.get(collect_logs, 'Unknown')} log collection enabled")
    report = None
    html_report_enabled = (mode != "scrape")
    try:
        if html_report_enabled:
            from core.reports_manager import ReportsManager, ENABLE_HTML, ENABLE_JSON
            if ENABLE_HTML or ENABLE_JSON:
                report = ReportsManager(flow_name=flow_name, run_id=run_id, output_dir=output_dir)
                report.start()
    except Exception as e:
        framework_logger.warning(f"HTML report initialization failed: {e}")

    flow_success = True
    flow_exception: Optional[BaseException] = None
    try:
        flow_func(stage_callback=stage_callback)
    except Exception as e:
        flow_success = False
        flow_exception = e
        framework_logger.error(f"Flow '{flow_name}' execution failed: {e}", exc_info=True)
    finally:
        if report:
            try:
                report.finalize(success=flow_success, exception=flow_exception)
            except Exception as rep_e:
                framework_logger.warning(f"Failed to finalize HTML report: {rep_e}")
        if not flow_success:
            raise SystemExit(1)

    if report and mode != "scrape":
        any_fail = any(s.status == 'FAIL' for s in report.steps)
        executed_any = any(s.status == 'PASS' for s in report.steps)
        has_not_executed = any(s.status == 'NOT_EXECUTED' for s in report.steps)
        interrupted = executed_any and has_not_executed and mode != 'noop'
        if any_fail or interrupted:
            reason_bits = []
            if any_fail:
                reason_bits.append('step failure detected')
            if interrupted:
                reason_bits.append('interrupted (partial execution)')
            framework_logger.error(f"Marking flow '{flow_name}' as FAILED: {', '.join(reason_bits)}")
            raise SystemExit(1)
 
    if mode != "noop":
        end_msg = f"{mode.title()} run for flow '{flow_name}'"
        if logs_enabled:
            end_msg += f" with logs for {lang_code} is completed. Results and Logs saved to: {output_dir}"
        else:
            end_msg += f" for {lang_code} is completed. Results saved to: {output_dir}"
    else:
        end_msg = f"No-op run for flow '{flow_name}'"
        if logs_enabled and output_dir:
            end_msg += f" with logs for {lang_code} is completed. Logs saved to: {output_dir}"
        else:
            end_msg += f" for {lang_code} is completed (no output generated)."
    framework_logger.info(end_msg)

def run_scrape_mode(flow_name: str, flow_args: Dict) -> None:
    """Helper to invoke scrape mode."""
    run_flow("scrape", flow_name, flow_args)

def run_validate_mode(flow_name: str, flow_args: Dict) -> None:
    """Helper to invoke validate mode."""
    run_flow("validate", flow_name, flow_args)

def run_noop_mode(flow_name: str, flow_args: Dict) -> None:
    """Helper to invoke noop mode."""
    run_flow("noop", flow_name, flow_args)

# ======================= Miscellaneous Utilities =======================

def load_json_file(file_path: str, encoding: str = "utf-8") -> dict:
    try:
        with open(file_path, "r", encoding=encoding) as f:
            return json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"JSON file not found: {file_path}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Error decoding JSON file {file_path}: {e}")

def load_locale_data() -> dict:
    return load_json_file(Constants.LOCALE_MAP_PATH)

def load_printer_profiles() -> dict:
    return load_json_file(Constants.PRINTER_PROFILES_PATH)

def load_address_for_country(country: str) -> dict:
    data = load_json_file(Constants.ADDRESS_DATA_PATH)
    return data.get(country)

def get_printer_details_for_country(
    profile_name: str,
    region: str,
    printer_profiles: dict,
    easy_enroll_flag: str = None,
    biz_model: str = "Flex"
) -> tuple:
    if profile_name not in printer_profiles:
        raise ValueError(f"Printer profile '{profile_name}' not found")
    skus_by_region = printer_profiles[profile_name].get("skus", {})
    if region not in skus_by_region:
        raise ValueError(f"No SKUs for region '{region}' in profile '{profile_name}'")
    skus = skus_by_region[region]
    if easy_enroll_flag is not None:
        want = easy_enroll_flag.lower() == "yes"
        skus = [s for s in skus if s.get("easy_enroll") == want]
        if not skus:
            raise ValueError(
                f"No SKUs with easy_enroll={want} for profile '{profile_name}', region '{region}'"
            )
    sku = skus[0]["sku"]
    supplyvariant = skus[0].get("supplyvariant")
    return profile_name, sku, supplyvariant, biz_model

def extract_svg_animation_duration(svg_xml, fallback_duration_ms=1000):
    import re
    match = re.search(r'dur="([\d\.]+)(ms|s)"', svg_xml)
    if match:
        val, unit = match.groups()
        seconds = float(val) / 1000 if unit == "ms" else float(val)
        return int(seconds * 1000)
    match_css = re.search(r'animation-duration:\s*([\d\.]+)(ms|s)', svg_xml)
    if match_css:
        val, unit = match_css.groups()
        seconds = float(val) / 1000 if unit == "ms" else float(val)
        return int(seconds * 1000)
    return fallback_duration_ms

def extract_animated_svgs(
    page,
    min_size=5120,
    poll_interval=0.02,
    max_frames=100,
    upscale_factor=4,
    save_debug_frames=True,
    debug_frame_dir="./debug_svg_frames"
):
    import io, time, os
    from PIL import Image, ImageChops
    from core.utils import hash_text

    results = []

    for idx, svg_elem in enumerate(page.query_selector_all("svg")):
        try:
            framework_logger.debug(f"SVG[{idx}] ---- Start capture ----")

            if not svg_elem.is_visible():
                framework_logger.debug(f"SVG[{idx}] skipped: not visible.")
                continue

            svg_xml = svg_elem.evaluate('el => el.outerHTML')
            size_bytes = len(svg_xml.encode("utf-8")) if svg_xml else 0
            framework_logger.debug(f"SVG[{idx}] outer_html size: {size_bytes}")
            if not svg_xml or size_bytes < min_size:
                framework_logger.debug(f"SVG[{idx}] skipped: too small or missing outer HTML.")
                continue

            # --- Animation detection (SMIL, CSS, JS) ---
            has_smil = svg_elem.evaluate(
                "el => !!el.querySelector('animate, animateTransform, animateMotion, animateColor, set')"
            )
            has_css = svg_elem.evaluate(
                '''
                el => {
                    const check = node => {
                        const style = window.getComputedStyle(node);
                        return (
                            (style.animationName && style.animationName !== 'none') ||
                            (style.transitionProperty && style.transitionProperty !== 'all' && style.transitionProperty !== 'none')
                        );
                    };
                    if (check(el)) return true;
                    for (const child of el.querySelectorAll('*')) {
                        if (check(child)) return true;
                    }
                    return false;
                }
                '''
            )
            has_timer = svg_elem.evaluate('el => !!window.setTimeout || !!window.setInterval')
            uses_raf = svg_elem.evaluate('el => !!window.requestAnimationFrame')
            has_script_tags = svg_elem.evaluate('el => !!el.querySelector("script")')
            script_contents = svg_elem.evaluate(
                'el => Array.from(el.querySelectorAll("script")).map(s=>s.textContent.slice(0, 150))'
            ) if has_script_tags else []

            has_event_handlers = svg_elem.evaluate(
                '''
                el => {
                    let handlers = [];
                    for (const evt of ["onload","onmousemove","onclick","onmousedown","onmouseup","onmouseover","onmouseout","onmouseenter","onmouseleave"]) {
                        if (el[evt]) handlers.push(evt);
                    }
                    for (const c of el.querySelectorAll("*")) {
                        for (const evt of ["onload","onmousemove","onclick","onmousedown","onmouseup","onmouseover","onmouseout","onmouseenter","onmouseleave"]) {
                            if (c[evt]) return true;
                        }
                    }
                    return handlers.length > 0;
                }
                '''
            )

            computed_style = svg_elem.evaluate(
                '''
                el => {
                    const style = window.getComputedStyle(el);
                    return {
                        animationName: style.animationName,
                        animationDuration: style.animationDuration,
                        transitionProperty: style.transitionProperty,
                        willChange: style.willChange,
                        opacity: style.opacity,
                        visibility: style.visibility,
                        display: style.display
                    };
                }
                '''
            )

            if has_smil and has_css:
                anim_type = "SMIL+CSS"
            elif has_smil:
                anim_type = "SMIL"
            elif has_css:
                anim_type = "CSS"
            else:
                anim_type = None

            is_animated = has_smil or has_css

            framework_logger.debug(
                f"SVG[{idx}] candidate (len={len(svg_xml)}): "
                f"SMIL={has_smil}, CSS={has_css}, AnimType={anim_type}, "
                f"JS_Timer={has_timer}, RAF={uses_raf}, Script={has_script_tags}, "
                f"ScriptContent={'YES' if script_contents else 'NO'}, "
                f"EventHandlers={has_event_handlers}, Style={computed_style}"
            )
            if script_contents:
                for snum, script in enumerate(script_contents):
                    framework_logger.debug(f"SVG[{idx}] script[{snum}] content[0:150]: {script!r}")

            # --- JS/CSS Animation Hard Reset for Determinism ---
            try:
                svg_elem.evaluate("""
                    el => {
                        const par = el.parentNode;
                        if (!par) return;
                        const next = el.nextSibling;
                        par.removeChild(el);
                        void document.body.offsetWidth;
                        if (next) par.insertBefore(el, next);
                        else par.appendChild(el);
                    }
                """)
                framework_logger.debug(f"SVG[{idx}] forcibly removed/re-added for hard reset.")
            except Exception as ex:
                framework_logger.debug(f"SVG[{idx}] hard reset via remove/add failed: {ex}")

            try:
                svg_elem.evaluate("""
                    el => {
                        el.classList.remove('animated');
                        void el.offsetWidth;
                        el.classList.add('animated');
                        el.style.animation = 'none';
                        el.querySelectorAll('*').forEach(child => {
                            child.style.animation = 'none';
                            void child.offsetWidth;
                            child.style.animation = '';
                        });
                        void el.offsetWidth;
                        el.style.animation = '';
                    }
                """)
                framework_logger.debug(f"SVG[{idx}] animation CSS class toggled and styles reset.")
            except Exception as ex:
                framework_logger.debug(f"SVG[{idx}] animation style reset failed: {ex}")

            try:
                svg_elem.hover()
                svg_elem.evaluate('el => el.dispatchEvent(new Event("mouseenter"))')
                framework_logger.debug(f"SVG[{idx}] hovered and mouseenter event dispatched")
            except Exception as ex:
                framework_logger.debug(f"SVG[{idx}] mouse event dispatch failed: {ex}")

            fixed_initial_delay = 0.20
            time.sleep(fixed_initial_delay)
            framework_logger.debug(f"SVG[{idx}] waited {fixed_initial_delay}s after trigger before first capture.")

            # --- Detect pixel-diffed animation (JS/Canvas driven) ---
            detect_frames = 5
            screenshots = []
            frame_hashes = []
            for i in range(detect_frames):
                t1 = time.time()
                svg_elem.scroll_into_view_if_needed()
                png_bytes = svg_elem.screenshot(timeout=5000)
                img = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
                screenshots.append(img)
                frame_bytes = img.tobytes()
                frame_hash = hash(frame_bytes)
                frame_hashes.append(frame_hash)
                t2 = time.time()
                framework_logger.debug(
                    f"SVG[{idx}] detect-frame {i} captured (shape={img.size}, {t2-t1:.4f}s)"
                )
                framework_logger.debug(f"SVG[{idx}] detect-frame {i} hash: {frame_hash}")

                if save_debug_frames:
                    if not os.path.exists(debug_frame_dir):
                        os.makedirs(debug_frame_dir, exist_ok=True)
                    img.save(f"{debug_frame_dir}/svg{idx}_detect_{i}.png")

                time.sleep(poll_interval)

            is_animated_pixel = False
            for i in range(1, detect_frames):
                diff = ImageChops.difference(screenshots[0], screenshots[i])
                diff_hist = diff.histogram()
                extrema = diff.getextrema()
                diff_sum = sum(diff_hist[1:])
                framework_logger.debug(
                    f"SVG[{idx}] diff frame 0 vs {i}: "
                    f"histogram_sum={diff_sum}, extrema={extrema}, bbox={diff.getbbox()}"
                )
                if diff.getbbox():
                    framework_logger.debug(f"SVG[{idx}] pixel-diff: frame 0 vs {i} is visually different (bbox != None)")
                    is_animated_pixel = True
                    break
                elif diff_sum > 0 or any(e != (0, 0) for e in extrema):
                    framework_logger.debug(f"SVG[{idx}] pixel-diff: frame 0 vs {i} detected (sensitive mode)")
                    is_animated_pixel = True
                    break
                elif frame_hashes[i] != frame_hashes[0]:
                    framework_logger.debug(
                        f"SVG[{idx}] frame 0 vs {i}: hashes differ but diff is 0 (possible browser re-encoding, alpha jitter, or invisible layer animation!)"
                    )

            if not is_animated and is_animated_pixel:
                framework_logger.debug(f"SVG[{idx}] flagged as animated by pixel-diff/extrema only.")
                is_animated = True

            if not is_animated:
                framework_logger.debug(f"SVG[{idx}] skipped: not animated (SMIL, CSS, or pixel-diff/extrema).")
                continue

            framework_logger.debug(f"SVG[{idx}] selected for GIF extraction (type={anim_type or 'pixel/extrema'}).")

            orig_width = svg_elem.evaluate("el => el.getAttribute('width')")
            orig_height = svg_elem.evaluate("el => el.getAttribute('height')")
            bbox = svg_elem.bounding_box()
            if bbox is None:
                framework_logger.debug(f"SVG[{idx}] skipped: bounding box is None.")
                continue
            target_width = int(bbox['width'] * upscale_factor)
            target_height = int(bbox['height'] * upscale_factor)
            svg_elem.evaluate(f"el => el.setAttribute('width', '{target_width}')")
            svg_elem.evaluate(f"el => el.setAttribute('height', '{target_height}')")
            framework_logger.debug(f"SVG[{idx}] upscaled for capture: {target_width}x{target_height} (orig: {orig_width}x{orig_height})")

            # --- Capture frames for GIF (detect actual animation speed) ---
            frames = []
            frame_hashes_gif = []
            times = []
            first_frame_hash = None
            loop_frame_idx = None
            for i in range(max_frames):
                now = svg_elem.evaluate("el => window.performance.now()")
                times.append(now)
                svg_elem.scroll_into_view_if_needed()
                png_bytes = svg_elem.screenshot(timeout=10000)
                img = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
                frame_bytes = img.tobytes()
                frame_hash = hash(frame_bytes)
                if i == 0:
                    first_frame_hash = frame_hash
                elif i > 2 and frame_hash == first_frame_hash:
                    loop_frame_idx = i
                    framework_logger.debug(f"SVG[{idx}] animation detected as looping at frame {i+1}.")
                    break
                frames.append(img)
                frame_hashes_gif.append(frame_hash)
                framework_logger.debug(f"SVG[{idx}] captured frame {i+1} (hash={frame_hash}, time={now})")
                time.sleep(poll_interval)

            # --- Calculate per-frame duration for GIF for visual speed sync ---
            per_frame_duration_ms = int(1000 * poll_interval)
            if loop_frame_idx is not None and loop_frame_idx > 0:
                loop_period = times[loop_frame_idx] - times[0]
                per_frame_duration_ms = max(1, int(loop_period / loop_frame_idx))
                framework_logger.debug(f"SVG[{idx}] loop period={loop_period:.2f}ms, frames={loop_frame_idx}, per_frame_duration={per_frame_duration_ms}ms")
            else:
                framework_logger.debug(f"SVG[{idx}] no loop detected, using poll_interval={poll_interval}s -> per_frame_duration={per_frame_duration_ms}ms")

            # Restore SVG size
            if orig_width:
                svg_elem.evaluate(f"el => el.setAttribute('width', '{orig_width}')")
            else:
                svg_elem.evaluate("el => el.removeAttribute('width')")
            if orig_height:
                svg_elem.evaluate(f"el => el.setAttribute('height', '{orig_height}')")
            else:
                svg_elem.evaluate("el => el.removeAttribute('height')")
            framework_logger.debug(f"SVG[{idx}] size restored after capture.")

            # --- In-memory GIF creation ---
            gif_bytes = None
            if len(frames) >= 2:
                palette_frames = [f.convert('P', palette=Image.ADAPTIVE, dither=Image.NONE) for f in frames]
                gif_io = io.BytesIO()
                palette_frames[0].save(
                    gif_io,
                    format='GIF',
                    save_all=True,
                    append_images=palette_frames[1:],
                    duration=per_frame_duration_ms,
                    loop=0,
                    optimize=True,
                    disposal=2
                )
                gif_bytes = gif_io.getvalue()
                framework_logger.debug(f"SVG[{idx}] GIF bytes generated ({len(frames)} frames, {target_width}x{target_height})")
            else:
                framework_logger.debug(f"SVG[{idx}] GIF not generated: <2 frames captured.")

            results.append({
                "element": svg_elem,
                "outer_html": svg_xml,
                "size": len(svg_xml.encode("utf-8")),
                "gif_bytes": gif_bytes,
                "frames_captured": len(frames),
                "resolution": (target_width, target_height),
                "animation_type": anim_type or ("pixel/extrema" if is_animated else "static"),
                "has_timer": has_timer,
                "uses_raf": uses_raf,
                "has_script_tags": has_script_tags,
                "event_handlers": has_event_handlers,
                "style": computed_style,
                "frame_hashes": frame_hashes_gif,
                "capture_times": times,
                "gif_frame_duration_ms": per_frame_duration_ms,
            })

        except Exception as e:
            framework_logger.debug(f"SVG[{idx}] extraction failed: {e}")
            continue
    return results


def extract_animated_gifs(page, min_size=1024):
    results = []
    for gif_elem in page.query_selector_all("img[src*='.gif']"):
        try:
            if not gif_elem.is_visible():
                continue
            src = gif_elem.get_attribute('src')
            if not src or '.gif' not in src.lower():
                continue
            response = page.context.request.get(src)
            if response.status != 200:
                continue
            gif_bytes = response.body()
            if not gif_bytes or len(gif_bytes) < min_size:
                continue
            try:
                with Image.open(io.BytesIO(gif_bytes)) as img:
                    is_animated = getattr(img, "is_animated", False) and img.n_frames > 1
            except Exception as e:
                is_animated = False
            if not is_animated:
                continue
            results.append({
                "element": gif_elem,
                "src": src,
                "bytes": gif_bytes,
                "size": len(gif_bytes)
            })
        except Exception as e:
            framework_logger.debug(f"GIF extraction failed: {e}")
            continue
    return results

def extract_animated_visuals(
    page,
    selector=".stepsImage",
    min_size=1024,
    screenshot_dir=".",
    context_hint=None,
    poll_interval=0.03,
    max_frames=60,
    detect_loop=True,
    save_debug_frames=False   # Default is False
):
    import os, io, time, hashlib
    from PIL import Image, ImageChops

    results = []
    idx = 0

    for elem in page.query_selector_all(selector):
        try:
            if not elem.is_visible():
                framework_logger.debug(f"VISUAL[{idx}] skipped: not visible.")
                continue

            html = elem.inner_html()
            tag = elem.evaluate('el => el.tagName')
            animation_type = None
            if '<svg' in html or elem.query_selector('svg'):
                animation_type = 'svg'
            elif elem.query_selector('canvas'):
                animation_type = 'canvas'
            elif "lottie" in html.lower():
                animation_type = 'lottie'
            else:
                style = elem.evaluate("el => window.getComputedStyle(el).backgroundImage")
                if style and "url(" in style:
                    animation_type = 'css_sprite'
                else:
                    animation_type = 'unknown'

            filename_hint = context_hint or animation_type or "visual"
            basename = f"visual_anim_{filename_hint}_{idx}"

            frames = []
            frame_hashes = []
            times = []
            loop_frame_idx = None

            for i in range(max_frames):
                # Force repaint for off-thread animations
                elem.evaluate("""
                    el => {
                        el.style.outline = ((Math.random()*100000)|0) + 'px solid transparent';
                        void el.offsetWidth;
                    }
                """)
                elem.scroll_into_view_if_needed()
                png_bytes = elem.screenshot()
                img = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
                frames.append(img)
                frame_hash = hashlib.sha256(img.tobytes()).hexdigest()
                frame_hashes.append(frame_hash)
                times.append(time.time())

                # Save per-frame PNGs ONLY if debugging is enabled
                if save_debug_frames:
                    os.makedirs(screenshot_dir, exist_ok=True)
                    img.save(os.path.join(screenshot_dir, f"{basename}_frame_{i}.png"))

                # Detect loop: if hashes start repeating
                if detect_loop and i > 3 and frame_hash == frame_hashes[0]:
                    loop_frame_idx = i
                    framework_logger.debug(f"VISUAL[{idx}] animation detected as looping at frame {i+1}.")
                    break

                time.sleep(poll_interval)

            # If there is more than one *unique* frame, treat as animation
            unique_hashes = set(frame_hashes)
            is_animated = len(unique_hashes) > 1

            gif_path = os.path.join(screenshot_dir, f"{basename}.gif")
            png_path = os.path.join(screenshot_dir, f"{basename}.png")
            out_path = png_path
            gif_bytes = None
            duration = int(1000 * poll_interval)

            if is_animated and len(frames) >= 2:
                if loop_frame_idx is not None and loop_frame_idx > 0:
                    total_period = times[loop_frame_idx] - times[0]
                    duration = max(1, int(total_period / loop_frame_idx * 1000))
                    end_idx = loop_frame_idx
                else:
                    end_idx = len(frames)
                palette_frames = [f.convert('P', palette=Image.ADAPTIVE, dither=Image.NONE) for f in frames[:end_idx]]
                palette_frames[0].save(
                    gif_path,
                    format='GIF',
                    save_all=True,
                    append_images=palette_frames[1:],
                    duration=duration,
                    loop=0,
                    optimize=True,
                    disposal=2
                )
                out_path = gif_path
                with open(gif_path, "rb") as f:
                    gif_bytes = f.read()
                framework_logger.debug(f"VISUAL[{idx}] animated GIF saved: {gif_path} ({end_idx} frames, duration/frame={duration}ms)")
            else:
                frames[0].save(png_path)
                with open(png_path, "rb") as f:
                    gif_bytes = f.read()
                framework_logger.debug(f"VISUAL[{idx}] static PNG saved: {png_path}")

            if len(gif_bytes) < min_size:
                os.remove(out_path)
                framework_logger.debug(f"VISUAL[{idx}] skipped: output < min_size ({len(gif_bytes)} bytes)")
                continue

            results.append({
                "element": elem,
                "animation_type": animation_type,
                "screenshot_path": out_path,
                "is_animated": is_animated,
                "frames_captured": len(frames),
                "duration_per_frame_ms": duration,
                "size": len(gif_bytes),
                "tag": tag
            })
            idx += 1

        except Exception as e:
            framework_logger.debug(f"VISUAL[{idx}] animation extraction failed: {e}")
            continue
    return results
