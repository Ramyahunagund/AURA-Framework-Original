# core/validator.py

import os
import hashlib
import ast
import shutil
import urllib.parse
import time 
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from PIL import Image
from playwright.sync_api import Page
from core.settings import framework_logger
from core.utils import (
    sanitize_filename,
    load_or_create_workbook,
    save_workbook,
    highlight_validation,
    save_full_page_screenshot,
    take_element_screenshot,
    take_element_handle_screenshot,
    hash_text,
    ensure_dirs,
    write_report_excel_context,
    extract_animated_svgs,
    extract_animated_gifs
)
from core.settings import Constants

COLOR_TOLERANCE = 35
MIN_TOLERANCE_PIXELS = 1

def parse_exclusions(exclude_str):
    if not exclude_str:
        return []
    try:
        exclusions = ast.literal_eval(exclude_str)
        if isinstance(exclusions, tuple):
            exclusions = [exclusions]
        return exclusions
    except Exception:
        return []

def is_in_exclusion(x, y, exclusions):
    for ex in exclusions:
        ex_x, ex_y, ex_w, ex_h = ex
        if ex_x <= x < ex_x + ex_w and ex_y <= y < ex_y + ex_h:
            return True
    return False

def get_bounding_boxes(pixels, size):
    if not pixels:
        return []
    width, height = size
    mask = Image.new("L", size, 0)
    for x, y in pixels:
        mask.putpixel((x, y), 255)
    visited = set()
    boxes = []
    def flood_fill(x, y):
        stack = [(x, y)]
        xmin = xmax = x
        ymin = ymax = y
        while stack:
            cx, cy = stack.pop()
            if (cx, cy) in visited or cx < 0 or cy < 0 or cx >= width or cy >= height:
                continue
            if mask.getpixel((cx, cy)) == 255:
                visited.add((cx, cy))
                xmin = min(xmin, cx)
                xmax = max(xmax, cx)
                ymin = min(ymin, cy)
                ymax = max(ymax, cy)
                stack.extend([(cx-1, cy), (cx+1, cy), (cx, cy-1), (cx, cy+1)])
        return (xmin, ymin, xmax - xmin + 1, ymax - ymin + 1) if xmax >= xmin and ymax >= ymin else None
    for y in range(height):
        for x in range(width):
            if mask.getpixel((x, y)) == 255 and (x, y) not in visited:
                box = flood_fill(x, y)
                if box and box[2] >= MIN_TOLERANCE_PIXELS and box[3] >= MIN_TOLERANCE_PIXELS:
                    boxes.append(box)
    return boxes

def compare_images_detect_then_filter(img1_path, img2_path, exclusions, color_tolerance=COLOR_TOLERANCE, original_exclude_str=None):
    if not os.path.exists(img1_path) or not os.path.exists(img2_path):
        return False, None, [], []
    if original_exclude_str is not None and original_exclude_str.strip() == '[]':
        return True, None, [], []
    img1 = Image.open(img1_path).convert("RGB")
    img2 = Image.open(img2_path).convert("RGB")
    w1, h1 = img1.size
    w2, h2 = img2.size

    width = max(w1, w2)
    height = max(h1, h2)

    base1 = Image.new("RGB", (width, height), (255, 255, 255))
    base2 = Image.new("RGB", (width, height), (255, 255, 255))
    base1.paste(img1, (0, 0))
    base2.paste(img2, (0, 0))

    failed_pixels = []
    failed_pixels_outside_exclusions = []

    for y in range(height):
        for x in range(width):
            rgb1 = base1.getpixel((x, y))
            rgb2 = base2.getpixel((x, y))
            diff = sum((c1 - c2) ** 2 for c1, c2 in zip(rgb1, rgb2)) ** 0.5
            if diff > color_tolerance:
                failed_pixels.append((x, y))
                if not is_in_exclusion(x, y, exclusions):
                    failed_pixels_outside_exclusions.append((x, y))

    boxes = get_bounding_boxes(failed_pixels_outside_exclusions, size=(width, height))
    if not boxes:
        return True, None, [], failed_pixels
    else:
        return False, None, boxes, failed_pixels_outside_exclusions

def overlay_true_failures(img2_path, failed_pixels_outside_exclusions, out_path, alpha=120):
    img2 = Image.open(img2_path).convert("RGBA")
    overlay = Image.new("RGBA", img2.size, (255, 0, 0, 0))
    overlay_data = overlay.load()
    width, height = img2.size
    for (x, y) in failed_pixels_outside_exclusions:
        if 0 <= x < width and 0 <= y < height:
            overlay_data[x, y] = (255, 0, 0, alpha)
    out_img = Image.alpha_composite(img2, overlay)
    out_img.save(out_path)
    return True

def compare_gif_files(gif_path_1: str, gif_path_2: str) -> bool:
    """
    Compare two GIF files byte-wise. Returns True if identical.
    """
    if not os.path.exists(gif_path_1) or not os.path.exists(gif_path_2):
        framework_logger.warning(f"GIF file(s) missing: {gif_path_1}, {gif_path_2}")
        return False
    with open(gif_path_1, 'rb') as f1, open(gif_path_2, 'rb') as f2:
        return f1.read() == f2.read()

def compare_svg_xml(svg_xml_1: str, svg_xml_2: str) -> bool:
    """
    Robust SVG comparison: ignores whitespace, attribute order, float precision, and comment nodes.
    Handles: attribute sorting, rounds floats, strips comments and normalizes text/tail.
    """
    import xml.etree.ElementTree as ET

    def normalize_float(val):
        # Try to round floats and float lists to 3 decimals
        try:
            return str(round(float(val), 3))
        except Exception:
            # try comma/space separated float lists (e.g. path data)
            parts = re.split(r'([ ,]+)', val)
            out = []
            for part in parts:
                try:
                    out.append(str(round(float(part), 3)))
                except Exception:
                    out.append(part.strip())
            return ''.join(out)

    def clean_element(elem):
        # Recursively sort attributes and children; clean text/tail; remove comments
        if elem is None or elem.tag is ET.Comment:
            return None

        attrib = {}
        for k, v in sorted(elem.attrib.items()):
            attrib[k] = normalize_float(v)
        elem.attrib.clear()
        elem.attrib.update(attrib)
        if elem.text: elem.text = elem.text.strip()
        if elem.tail: elem.tail = elem.tail.strip()
        # Remove comments
        children = list(elem)
        for child in children:
            if ET.iselement(child) and (child.tag is ET.Comment or str(child.tag).lower() == 'comment'):
                elem.remove(child)
            else:
                clean_element(child)
        # Sort children by tag+attrib+text for deterministic order
        elem[:] = sorted(elem, key=lambda e: (e.tag, str(sorted(e.attrib.items())), e.text or ""))
        return elem

    def get_root(xml: str):
        try:
            return ET.fromstring(xml)
        except Exception as ex:
            framework_logger.warning(f"[SVG_COMPARE] Parse error: {ex}")
            return None

    root1, root2 = get_root(svg_xml_1), get_root(svg_xml_2)
    if root1 is None or root2 is None:
        # Fallback to whitespace-insensitive
        return re.sub(r'\s+', '', svg_xml_1) == re.sub(r'\s+', '', svg_xml_2)

    clean_element(root1)
    clean_element(root2)
    # Serialize back as string
    norm1 = ET.tostring(root1, encoding="unicode")
    norm2 = ET.tostring(root2, encoding="unicode")
    return norm1 == norm2


def validate_context(flow_name, context_name, page, excel_path, output_folder, validation_workbook_path=None, screenshot_only=False, sub_context=None, lang_code=None, animations=False, target_dir=None):
    val_ss_dir = os.path.join(output_folder, "validation_screenshots")
    fail_ss_dir = os.path.join(output_folder, "validation_failure_screenshots")
    val_anim_dir = os.path.join(output_folder, "validation_animations")
    fail_anim_dir = os.path.join(output_folder, "validation_failure_animations")
    ensure_dirs(val_ss_dir, fail_ss_dir, val_anim_dir, fail_anim_dir)

    subctx_suffix = f"_{sub_context}" if sub_context else ""

    if screenshot_only:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if lang_code:
            validation_fname = f"{context_name}{subctx_suffix}_{lang_code}_full_page.png"
        else:
            validation_fname = f"{context_name}{subctx_suffix}_full_page.png"
        validation_path = os.path.join(val_ss_dir, validation_fname)
        # page.screenshot(path=validation_path, full_page=True)
        for attempt in range(3):
            try:
                page.screenshot(path=validation_path, full_page=True, timeout=60000)
                break
            except Exception as ex:
                if attempt == 2:
                    framework_logger.error(f"[VALIDATE] Screenshot failed after retries: {ex}")
                    raise
                time.sleep(2)
        framework_logger.info(f"[VALIDATOR] Screenshot-only mode: Saved full-page screenshot for {flow_name}/{context_name}{subctx_suffix} at {validation_path}")

        header = [
            "Context", "SubContext", "PageURL", "Timestamp", "Type", "Tag",
            "Text_Expected", "Text_Actual",
            "Href_Expected", "Href_Actual",
            "Text_Result", "Screenshot_Baseline", "Screenshot_Validation",
            "Screenshot_Result", "Diff_Image", "ExcludeRegions", "SuggestedExcludeRegions", "Comments"
        ]

        scraped_wb = load_workbook(excel_path)
        ws = scraped_wb[context_name] if context_name in scraped_wb.sheetnames else scraped_wb.active
        master_header = [cell.value for cell in ws[1]]
        sub_context_idx = master_header.index("SubContext") if "SubContext" in master_header else None

        found_baseline = False
        report_row = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            item = dict(zip(master_header, row))
            context_match = item.get("Context", "") == context_name or ws.title == context_name
            subcontext_match = True
            if sub_context_idx is not None:
                item_subctx = row[sub_context_idx]
                subcontext_match = (item_subctx or "") == (sub_context or "")
            is_fullpage = ((item.get("Type", "") or "").lower() == "fullpage" or str(item.get("Screenshot", "")).endswith("full_page.png"))
            if context_match and subcontext_match and is_fullpage:
                baseline_screenshot = item.get("Screenshot", "")
                baseline_img_path = os.path.join(os.path.dirname(excel_path), "screenshots", baseline_screenshot)
                exclusions = parse_exclusions(item.get("ExcludeRegions", ""))
                exclude_str = item.get("ExcludeRegions", "")
                is_pass, diff, suggested_boxes, failed_pixels_outside_exclusions = compare_images_detect_then_filter(
                    baseline_img_path, validation_path, exclusions, color_tolerance=COLOR_TOLERANCE, original_exclude_str=exclude_str)
                report_row = {
                    "Context": context_name,
                    "SubContext": sub_context or "",
                    "PageURL": page.url,
                    "Timestamp": timestamp,
                    "Type": "FullPage",
                    "Tag": "",
                    "Text_Expected": "",
                    "Text_Actual": "",
                    "Href_Expected": "",
                    "Href_Actual": "",
                    "Text_Result": "",
                    "Screenshot_Baseline": baseline_screenshot,
                    "Screenshot_Validation": f"{context_name}{subctx_suffix}_full_page.png",
                    "Screenshot_Result": "PASS" if is_pass else "FAIL",
                    "Diff_Image": "",
                    "ExcludeRegions": str(exclusions) if exclusions else "",
                    "SuggestedExcludeRegions": str(suggested_boxes) if not is_pass else "",
                    "Comments": "",
                }
                if is_pass:
                    if failed_pixels_outside_exclusions:
                        report_row["Screenshot_Result"] = "PASS_WITH_EXCLUDES"
                        report_row["Comments"] += "All visual differences are within ExcludeRegions."
                    else:
                        report_row["Screenshot_Result"] = "PASS"
                else:
                    diff_fname = f"fail_0000_{context_name}{subctx_suffix}_full_page.png"
                    diff_path = os.path.join(fail_ss_dir, diff_fname)
                    overlayed = overlay_true_failures(validation_path, failed_pixels_outside_exclusions, diff_path)
                    if overlayed:
                        report_row["Diff_Image"] = diff_fname
                        report_row["Comments"] += "Red overlay shows differences (see image)."
                    else:
                        report_row["Comments"] += "Images differ but diff could not be created."
                found_baseline = True
                break

        if not found_baseline:
            report_row = {
                "Context": context_name,
                "SubContext": sub_context or "",
                "PageURL": page.url,
                "Timestamp": timestamp,
                "Type": "FullPage",
                "Tag": "",
                "Text_Expected": "",
                "Text_Actual": "",
                "Href_Expected": "",
                "Href_Actual": "",
                "Text_Result": "",
                "Screenshot_Baseline": "",
                "Screenshot_Validation": f"{context_name}{subctx_suffix}_full_page.png",
                "Screenshot_Result": "N/A",
                "Diff_Image": "",
                "ExcludeRegions": "",
                "SuggestedExcludeRegions": "",
                "Comments": "No FullPage baseline found for comparison."
            }

        if validation_workbook_path is None:
            validation_workbook_path = os.path.join(output_folder, "validation_results.xlsx")
        wb, _ = load_or_create_workbook(validation_workbook_path, Constants.VALIDATION_HEADERS)
        write_report_excel_context([report_row], wb, context_name, validation_workbook_path)
        save_workbook(wb, validation_workbook_path)
        framework_logger.info(f"[VALIDATOR] Validation report for context '{context_name}{subctx_suffix}' written to {validation_workbook_path}")
        return

    if validation_workbook_path is None:
        validation_workbook_path = os.path.join(output_folder, "validation_results.xlsx")
    wb, _ = load_or_create_workbook(validation_workbook_path, Constants.VALIDATION_HEADERS)

    scraped_wb = load_workbook(excel_path)
    if context_name in scraped_wb.sheetnames:
        ws = scraped_wb[context_name]
    else:
        ws = scraped_wb.active
    header = None
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if row and any(cell is not None and str(cell).strip() != '' for cell in row):
            header = [str(cell).strip() if cell is not None else '' for cell in row]
            break
    if header is None:
        raise ValueError(f"No header row found in sheet '{ws.title}' of {excel_path}")
    sub_context_idx = None
    if "SubContext" in header:
        sub_context_idx = header.index("SubContext")

    master_data = []
    for row in ws.iter_rows(min_row=ws.min_row+1, values_only=True):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        item = dict(zip(header, row))
        context_match = item.get("Context", "") == context_name or ws.title == context_name
        subcontext_match = True
        if sub_context_idx is not None:
            item_subctx = row[sub_context_idx]
            subcontext_match = (item_subctx or "") == (sub_context or "")
        if context_match and subcontext_match:
            master_data.append(item)

    results = []
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # ---- Extract all SVGs into validation_animations folder ONCE (if animations enabled) ----
    if animations:
        svg_animations = extract_animated_svgs(page)
        for svg_info in svg_animations:
            svg_xml = svg_info["outer_html"]
            svg_hash = hash_text(svg_xml)
            svg_filename = f"{context_name}{subctx_suffix}_svg_{svg_hash}.svg"
            svg_candidate_path = os.path.join(val_anim_dir, svg_filename)
            # Always overwrite
            with open(svg_candidate_path, "w", encoding="utf-8") as f:
                f.write(svg_xml)
            framework_logger.info(f"[VALIDATOR] Saved SVG: {svg_candidate_path}")

    handles = page.query_selector_all("body *")
    live_elements = []
    for idx, handle in enumerate(handles):
        try:
            if not handle.is_visible():
                continue
            text = handle.inner_text().strip()
            if not text:
                continue
            box = handle.bounding_box()
            if not box or box['width'] == 0 or box['height'] == 0:
                continue
            tag = handle.evaluate("el => el.tagName")
            href = handle.get_attribute("href") or ""
            onclick = handle.get_attribute("onclick") or ""
            live_elements.append({
                "tag": tag,
                "text": text,
                "href": href,
                "onclick": onclick,
                "handle": handle
            })
        except Exception as ex:
            if "Node is not an HTMLElement" in str(ex):
                continue
            framework_logger.warning(f"[VALIDATOR] Error processing element: {ex}")
            continue

    for idx, master in enumerate(master_data):
        master_type = (master.get("Type", "") or "").lower()
        master_screenshot = master.get("Screenshot", "") or ""
        exclusions = parse_exclusions(master.get("ExcludeRegions", ""))
        exclude_str = master.get("ExcludeRegions", "")
        master_subctx = master.get("SubContext", "") if "SubContext" in master else ""

        # Animation validation gated by 'animations' flag
        if master_type == "animation":
            if not animations:
                continue
            report_row = {
                "Context": context_name,
                "SubContext": master_subctx,
                "PageURL": page.url,
                "Timestamp": timestamp,
                "Type": master.get("Type", ""),
                "Tag": master.get("Tag", ""),
                "Text_Expected": "",
                "Text_Actual": "",
                "Href_Expected": master.get("Href", ""),
                "Href_Actual": "",
                "Text_Result": "",
                "Screenshot_Baseline": master.get("Screenshot", ""),
                "Screenshot_Validation": "",
                "Screenshot_Result": "",
                "Diff_Image": "",
                "ExcludeRegions": str(exclusions) if exclusions else "",
                "SuggestedExcludeRegions": "",
                "Comments": ""
            }
            baseline_file = master.get("Screenshot", "")
            baseline_href = master.get("Href", "")  # src for GIF, empty for SVG
            baseline_path = os.path.join(os.path.dirname(excel_path), "animations", baseline_file)
            validation_path = os.path.join(val_anim_dir, baseline_file)
            tag = (master.get("Tag", "") or "").lower()
            found = False

            try:
                os.makedirs(os.path.dirname(validation_path), exist_ok=True)
                if tag == "gif":
                    # Extract all animated GIFs centrally
                    gif_animations = extract_animated_gifs(page)
                    for gif_info in gif_animations:
                        gif_bytes = gif_info["bytes"]
                        gif_hash = hash_text(gif_bytes)
                        gif_filename = f"{context_name}{subctx_suffix}_gif_{gif_hash}.gif"
                        if gif_filename == baseline_file:
                            with open(validation_path, 'wb') as f:
                                f.write(gif_bytes)
                            framework_logger.info(f"[VALIDATOR] Saved validation GIF: {validation_path}")
                            found = True
                            break
                    if not found:
                        framework_logger.warning(f"[VALIDATOR] Could not find matching GIF for validation: {baseline_file}")

                elif tag == "svg":
                    if not os.path.exists(baseline_path):
                        report_row["Screenshot_Result"] = "N/A"
                        report_row["Comments"] = f"Baseline SVG not found: {baseline_path}"
                        results.append(report_row)
                        continue

                    if not os.path.exists(validation_path):
                        report_row["Screenshot_Result"] = "N/A"
                        report_row["Comments"] = f"Validation SVG not found: {validation_path}"
                        results.append(report_row)
                        continue

                    with open(baseline_path, "r", encoding="utf-8") as f:
                        ref_svg_xml = f.read()
                    with open(validation_path, "r", encoding="utf-8") as f:
                        cap_svg_xml = f.read()
                    is_pass = compare_svg_xml(ref_svg_xml, cap_svg_xml)
                    if is_pass:
                        report_row["Screenshot_Validation"] = baseline_file
                        report_row["Screenshot_Result"] = "PASS"
                        report_row["Comments"] = "SVG animation match."
                    else:
                        report_row["Screenshot_Validation"] = baseline_file
                        report_row["Screenshot_Result"] = "FAIL"
                        report_row["Comments"] = "SVG animation mismatch."
                        fail_path = os.path.join(fail_anim_dir, f"fail_{baseline_file}")
                        shutil.copy(validation_path, fail_path)
                        report_row["Diff_Image"] = os.path.relpath(fail_path, output_folder)
                    results.append(report_row)
                    continue

                else:
                    report_row["Screenshot_Result"] = "N/A"
                    report_row["Comments"] = f"Unknown animation tag: {tag}"

                # Now perform the comparison and copy failure asset if needed:
                if found and tag == "gif":
                    if os.path.exists(baseline_path) and os.path.exists(validation_path):
                        is_pass = compare_gif_files(baseline_path, validation_path)
                        if is_pass:
                            report_row["Screenshot_Result"] = "PASS"
                            report_row["Comments"] = "GIF animation match."
                        else:
                            report_row["Screenshot_Result"] = "FAIL"
                            report_row["Comments"] = "GIF animation mismatch."
                            fail_path = os.path.join(fail_anim_dir, f"fail_{baseline_file}")
                            shutil.copy(validation_path, fail_path)
                            report_row["Diff_Image"] = os.path.relpath(fail_path, output_folder)
                    else:
                        report_row["Screenshot_Result"] = "N/A"
                        report_row["Comments"] = f"GIF file missing for comparison. Baseline exists: {os.path.exists(baseline_path)}, Validation exists: {os.path.exists(validation_path)}"

                elif found and tag == "svg":
                    if os.path.exists(baseline_path) and os.path.exists(validation_path):
                        with open(baseline_path, "r", encoding="utf-8") as f:
                            ref_svg_xml = f.read()
                        with open(validation_path, "r", encoding="utf-8") as f:
                            cap_svg_xml = f.read()
                        is_pass = compare_svg_xml(ref_svg_xml, cap_svg_xml)
                        if is_pass:
                            report_row["Screenshot_Result"] = "PASS"
                            report_row["Comments"] = "SVG animation match."
                        else:
                            report_row["Screenshot_Result"] = "FAIL"
                            report_row["Comments"] = "SVG animation mismatch."
                            fail_path = os.path.join(fail_anim_dir, f"fail_{baseline_file}")
                            shutil.copy(validation_path, fail_path)
                            report_row["Diff_Image"] = os.path.relpath(fail_path, output_folder)
                    else:
                        report_row["Screenshot_Result"] = "N/A"
                        report_row["Comments"] = f"SVG file missing for comparison. Baseline exists: {os.path.exists(baseline_path)}, Validation exists: {os.path.exists(validation_path)}"

            except Exception as e:
                framework_logger.warning(f"[VALIDATOR] Error capturing animation asset for validation: {e}")

            results.append(report_row)
            continue


        if master_type == "fullpage" or master_screenshot.endswith("full_page.png"):
            report_row = {
                "Context": context_name,
                "SubContext": master_subctx,
                "PageURL": page.url,
                "Timestamp": timestamp,
                "Type": master.get("Type", ""),
                "Tag": "",
                "Text_Expected": "",
                "Text_Actual": "",
                "Href_Expected": "",
                "Href_Actual": "",
                "Text_Result": "",
                "Screenshot_Baseline": master_screenshot,
                "Screenshot_Validation": f"{context_name}{subctx_suffix}_full_page.png",
                "Screenshot_Result": "",
                "Diff_Image": "",
                "ExcludeRegions": str(exclusions) if exclusions else "",
                "SuggestedExcludeRegions": "",
                "Comments": ""
            }
            validation_path = os.path.join(val_ss_dir, f"{context_name}{subctx_suffix}_full_page.png")
            page.screenshot(path=validation_path, full_page=True)
            baseline_img_path = os.path.join(os.path.dirname(excel_path), "screenshots", master_screenshot)
            is_pass, diff, suggested_boxes, failed_pixels_outside_exclusions = compare_images_detect_then_filter(
                baseline_img_path, validation_path, exclusions, color_tolerance=COLOR_TOLERANCE, original_exclude_str=exclude_str)
            report_row["Screenshot_Result"] = "PASS" if is_pass else "FAIL"
            if is_pass:
                if failed_pixels_outside_exclusions:
                    report_row["Screenshot_Result"] = "PASS_WITH_EXCLUDES"
                    report_row["Diff_Image"] = ""
                    report_row["SuggestedExcludeRegions"] = ""
                    report_row["Comments"] += "All visual differences are within ExcludeRegions."
                else:
                    report_row["Screenshot_Result"] = "PASS"
                    report_row["Diff_Image"] = ""
                    report_row["SuggestedExcludeRegions"] = ""
            else:
                diff_fname = f"fail_{idx:04d}_{context_name}{subctx_suffix}_full_page.png"
                diff_path = os.path.join(fail_ss_dir, diff_fname)
                overlayed = overlay_true_failures(validation_path, failed_pixels_outside_exclusions, diff_path)
                if overlayed:
                    report_row["Diff_Image"] = diff_fname
                    report_row["SuggestedExcludeRegions"] = str(suggested_boxes)
                    report_row["Comments"] += "Red overlay shows differences (see image)."
                else:
                    report_row["Diff_Image"] = ""
                    report_row["Comments"] += "Images differ but diff could not be created."
            results.append(report_row)
            continue

        master_text = (master.get("Text") or "").strip()
        master_tag = (master.get("Tag") or "").strip()
        master_href = (master.get("Href") or "").strip()
        found = False
        match_idx = None
        for j, live in enumerate(live_elements):
            if (master_text == (live["text"] or "").strip()) and (master_tag == live["tag"]):
                if master_href and master_href != live["href"]:
                    continue
                found = True
                match_idx = j
                break
        report_row = {
            "Context": context_name,
            "SubContext": master_subctx,
            "PageURL": page.url,
            "Timestamp": timestamp,
            "Type": master.get("Type", ""),
            "Tag": master.get("Tag", ""),
            "Text_Expected": master_text,
            "Text_Actual": "",
            "Href_Expected": master_href,
            "Href_Actual": "",
            "Text_Result": "FAIL",
            "Screenshot_Baseline": master_screenshot,
            "Screenshot_Validation": "",
            "Screenshot_Result": "N/A",
            "Diff_Image": "",
            "ExcludeRegions": str(exclusions) if exclusions else "",
            "SuggestedExcludeRegions": "",
            "Comments": ""
        }
        if found and match_idx is not None:
            live = live_elements.pop(match_idx)
            report_row["Text_Actual"] = live["text"]
            report_row["Href_Actual"] = live["href"]
            report_row["Text_Result"] = "PASS" if master_text == live["text"] and (not master_href or master_href == live["href"]) else "FAIL"

            if master_screenshot:
                validation_fname = f"validation_{idx:04d}_{context_name}{subctx_suffix}_{lang_code}_{hash_text(live['text'])}.png"
                validation_path = os.path.join(val_ss_dir, validation_fname)
                try:
                    shot_ok = take_element_handle_screenshot(live["handle"], validation_path)
                    if shot_ok:
                        report_row["Screenshot_Validation"] = validation_fname
                        baseline_img_path = os.path.join(os.path.dirname(excel_path), "screenshots", master_screenshot)
                        is_pass, diff, suggested_boxes, failed_pixels_outside_exclusions = compare_images_detect_then_filter(
                            baseline_img_path, validation_path, exclusions, color_tolerance=COLOR_TOLERANCE, original_exclude_str=exclude_str)
                        if is_pass:
                            if failed_pixels_outside_exclusions:
                                report_row["Screenshot_Result"] = "PASS_WITH_EXCLUDES"
                                report_row["Diff_Image"] = ""
                                report_row["SuggestedExcludeRegions"] = ""
                                report_row["Comments"] += "All visual differences are within ExcludeRegions."
                            else:
                                report_row["Screenshot_Result"] = "PASS"
                                report_row["Diff_Image"] = ""
                                report_row["SuggestedExcludeRegions"] = ""
                        else:
                            report_row["Screenshot_Result"] = "FAIL"
                            diff_fname = f"fail_{idx:04d}_{context_name}{subctx_suffix}_{lang_code}_{hash_text(live['text'])}.png"
                            diff_path = os.path.join(fail_ss_dir, diff_fname)
                            overlayed = overlay_true_failures(validation_path, failed_pixels_outside_exclusions, diff_path)
                            if overlayed:
                                report_row["Diff_Image"] = diff_fname
                                report_row["SuggestedExcludeRegions"] = str(suggested_boxes)
                                report_row["Comments"] += "Red overlay shows differences (see image)."
                            else:
                                report_row["Diff_Image"] = ""
                                report_row["Comments"] += "Images differ but diff could not be created."
                except Exception as ex:
                    report_row["Screenshot_Validation"] = ""
                    report_row["Screenshot_Result"] = "ERROR"
                    report_row["Comments"] += f" Screenshot error: {ex}"
            else:
                report_row["Screenshot_Validation"] = ""
                report_row["Screenshot_Result"] = "N/A"
                report_row["Diff_Image"] = ""
                report_row["SuggestedExcludeRegions"] = ""
                if report_row["Text_Result"] == "PASS":
                    report_row["Comments"] += "Text validated only (no screenshot baseline provided)."
                else:
                    report_row["Comments"] += "Text mismatch and no screenshot baseline provided."
        else:
            report_row["Comments"] = "Element not found on live page"
        results.append(report_row)

    write_report_excel_context(results, wb, context_name, validation_workbook_path)
    wb.save(validation_workbook_path)
    framework_logger.info(f"[VALIDATOR] Validation report for context '{context_name}{subctx_suffix}' written to {validation_workbook_path}")
